import pandas as pd
import openpyxl

def read_maze(filename):
    dataFrame = openpyxl.load_workbook(filename)

    dataFrame = dataFrame.active

    print(dataFrame)

    numberRows = dataFrame.max_row
    numberCols = dataFrame.max_column

    maze = [[0 for x in range(numberRows)] for y in range(numberCols)]

    for row in range(0, numberRows):
        for col, column in zip(dataFrame.iter_cols(0, dataFrame.max_column), range(0, numberCols)):
            maze[row][column] = col[row].value
    return maze

def is_validToMove(maze, row, col):
    if row >= 0 and row < len(maze) and col >= 0 and col < len(maze[0]) and maze[row][col] != -1:
        return True
    else:
        return False
def find_paths(maze, row, col, path, cost, paths):
    if row == len(maze) - 1 and col == len(maze[0]) - 1:

        paths.append((list(path), cost))
        return

    if is_validToMove(maze, row+1, col):
        path.append((row+1, col))
        find_paths(maze, row+1, col, path, cost+maze[row+1][col], paths)
        path.pop()

    if is_validToMove(maze, row, col+1):
        path.append((row, col+1))
        find_paths(maze, row, col+1, path, cost+maze[row][col+1], paths)
        path.pop()

def display_paths(paths):
    print("All possible paths:")
    for path, cost in paths:
        print(path, "with cost", cost)
        
def find_lowest_cost_path(paths):
    lowest_cost = 10000000000000
    lowest_cost_path = None
    print(paths)
    for path, cost in paths:
        if cost < lowest_cost:
            lowest_cost = cost
            lowest_cost_path = path
    return lowest_cost_path, lowest_cost

def main():
    filename = input("Enter the filename of the maze: ")
    maze = read_maze(filename)

    paths = []
    find_paths(maze, 0, 0, [(0, 0)], maze[0][0], paths)

    display_paths(paths)

    return paths

def get_lowest_cost_path(paths):
    if not paths:
        return None
    lowest_cost_path = paths[0]
    for path in paths:
        if path[1] < lowest_cost_path[1]:
            lowest_cost_path = path
    return lowest_cost_path

paths = main()

print("Lowest Cost Path" + str(get_lowest_cost_path(paths)))
